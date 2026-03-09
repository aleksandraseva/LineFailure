from openpyxl import load_workbook
import os
import re

from Lines.models.Connection import Connection


def find_connection():
    try:
        j = 0
        print(os.path)
        file = os.path.join("Konekcije.xlsx")
        wb = load_workbook(file, read_only=True)
        sheet = wb.active

        pattern = re.compile(
            r"(unit-[^\s]+\s+port-[^\s]+.*?)(?=unit-[^\s]+\s+port-[^\s]+|$)", re.DOTALL
        )
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            red_str = " ".join(map(str, filter(None, row)))
            blocks = pattern.findall(red_str)
            for block in blocks:
                find_data = r"^(?:(?P<unit>unit-\d+)\s+(?P<port>port-\d+)\s+(?P<locations>.+)\s+(?P<port2>port-\d+)\s+(?P<unit2>unit-\d+))"

                m = re.match(find_data, block)
                if m:
                    unit1 = m.group("unit")
                    port1 = m.group("port")
                    locations = m.group("locations")
                    port2 = m.group("port2")
                    unit2 = m.group("unit2")

                    find_locations = re.compile(
                        r"/([A-Za-z0-9]+(?:\s*\([^)]+\))?)", re.DOTALL
                    )
                    data = find_locations.findall(locations)
                    location1 = data[0]
                    location2 = data[1] // ovdje ima red gdje nema

                    print(j)

                    Connection(
                        first_location=location1,
                        first_unit=unit1,
                        first_port=port1,
                        second_location=location2,
                        second_unit=unit2,
                        second_port=port2,
                    ).save()
                    j = j + 1

    except Exception as e:
        print(e)


if __name__ == "__main__":
    find_connection()
